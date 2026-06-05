from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd

from contaclass.models.entry import HistoricalEntry, NewEntry

COLUMN_ALIASES = {
    "data": "entry_date",
    "date": "entry_date",
    "dt": "entry_date",
    "lançamento": "entry_date",
    "lancamento": "entry_date",
    "cod débito": "debit_code",
    "cod debito": "debit_code",
    "codigo debito": "debit_code",
    "código débito": "debit_code",
    "codigo débito": "debit_code",
    "debito": "debit_code",
    "débito": "debit_code",
    "cd": "debit_code",
    "cod crédito": "credit_code",
    "cod credito": "credit_code",
    "codigo credito": "credit_code",
    "código crédito": "credit_code",
    "codigo crédito": "credit_code",
    "credito": "credit_code",
    "crédito": "credit_code",
    "cc": "credit_code",
    "valor": "amount",
    "value": "amount",
    "val": "amount",
    "vlr": "amount",
    "fornecedor": "supplier",
    "fornecedores": "supplier",
    "favorecido": "supplier",
    "beneficiario": "supplier",
    "beneficiário": "supplier",
    "descricao": "supplier",
    "descrição": "supplier",
    "desc": "supplier",
    "nome": "supplier",
    "historico": "supplier",
    "histórico": "supplier",
}


class ColumnDetectionError(Exception):
    pass


class ExcelReader:
    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)

    def list_tabs(self) -> list[dict]:
        wb = openpyxl.load_workbook(self.filepath, read_only=True)
        tabs = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tabs.append({
                "name": sheet_name,
                "rows": ws.max_row or 0,
            })
        wb.close()
        return tabs

    def read_historical(
        self, tab_filter: list[str] | None = None, column_map: dict[str, str] | None = None
    ) -> list[HistoricalEntry]:
        xls = pd.ExcelFile(self.filepath)
        entries = []

        for sheet_name in xls.sheet_names:
            if tab_filter and sheet_name not in tab_filter:
                continue

            df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
            if len(df) < 2:
                continue

            clean = self._clean_dataframe(df)
            mapping = self._detect_columns(list(clean.columns), column_map)

            for _, row in clean.iterrows():
                try:
                    entry_date = self._parse_date(row.get(mapping["entry_date"]))
                except (ValueError, TypeError):
                    continue

                try:
                    amount = self._parse_amount(row.get(mapping["amount"]))
                except (ValueError, TypeError):
                    continue

                raw_supplier = str(row.get(mapping["supplier"], "")).strip()
                if not raw_supplier or raw_supplier in ("nan", "None", ""):
                    continue

                debit = str(row.get(mapping["debit_code"], ""))
                credit = str(row.get(mapping["credit_code"], ""))

                if debit in ("nan", "None"):
                    debit = ""
                if credit in ("nan", "None"):
                    credit = ""

                entries.append(HistoricalEntry(
                    tab_name=sheet_name,
                    entry_date=entry_date,
                    raw_supplier=raw_supplier,
                    normalized_supplier="",
                    debit_code=debit,
                    credit_code=credit,
                    amount=amount,
                    row_number=len(entries) + 1,
                ))

        xls.close()
        return entries

    def read_new(
        self, tab_name: str | None = None, column_map: dict[str, str] | None = None
    ) -> list[NewEntry]:
        xls = pd.ExcelFile(self.filepath)
        all_entries = []

        for sheet_name in xls.sheet_names:
            if tab_name and sheet_name != tab_name:
                continue

            df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
            if len(df) < 2:
                continue

            clean = self._clean_dataframe(df)
            mapping = self._detect_new_columns(list(clean.columns), column_map)

            for idx, row in clean.iterrows():
                try:
                    entry_date = self._parse_date(row.get(mapping["entry_date"]))
                except (ValueError, TypeError):
                    continue

                try:
                    amount = self._parse_amount(row.get(mapping["amount"]))
                except (ValueError, TypeError):
                    continue

                raw_supplier = str(row.get(mapping["supplier"], "")).strip()
                if not raw_supplier or raw_supplier in ("nan", "None", ""):
                    continue

                all_entries.append(NewEntry(
                    row_number=len(all_entries) + 1,
                    entry_date=entry_date,
                    raw_supplier=raw_supplier,
                    amount=amount,
                ))

        xls.close()
        return all_entries

    def preview_new(self, n_rows: int = 10) -> list[dict]:
        df = pd.read_excel(self.filepath, header=0, nrows=n_rows)
        mapping = self._detect_new_columns(list(df.columns))
        preview = []
        for _, row in df.head(n_rows).iterrows():
            preview.append({
                "entry_date": str(row.get(mapping["entry_date"], "")),
                "amount": str(row.get(mapping["amount"], "")),
                "supplier": str(row.get(mapping["supplier"], "")),
            })
        return preview

    def _detect_columns(
        self, columns: list[str], manual_map: dict[str, str] | None = None
    ) -> dict[str, str]:
        if manual_map:
            return manual_map

        result: dict[str, str | None] = {
            "entry_date": None,
            "debit_code": None,
            "credit_code": None,
            "amount": None,
            "supplier": None,
        }

        for col in columns:
            clean = col.strip().lower()
            if clean in COLUMN_ALIASES:
                target = COLUMN_ALIASES[clean]
                result[target] = col

        missing = [k for k, v in result.items() if v is None]
        if missing:
            result = self._infer_by_position(columns, result)

        return result

    def _detect_new_columns(
        self, columns: list[str], manual_map: dict[str, str] | None = None
    ) -> dict[str, str]:
        if manual_map:
            return manual_map

        result: dict[str, str | None] = {
            "entry_date": None,
            "amount": None,
            "supplier": None,
        }

        for col in columns:
            clean = col.strip().lower()
            if clean in COLUMN_ALIASES:
                target = COLUMN_ALIASES[clean]
                if target in result:
                    result[target] = col

        missing = [k for k, v in result.items() if v is None]
        if missing:
            result = self._infer_by_position(columns, result)

        return result

    def _infer_by_position(self, columns: list[str], current: dict) -> dict:
        inferred = dict(current)
        for key, val in current.items():
            if val is not None:
                continue

            idx = 0 if key == "entry_date" else (1 if key == "amount" else (2 if key == "supplier" else 3))
            if idx < len(columns):
                inferred[key] = columns[idx]

        return inferred

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df.columns = [str(c).strip() for c in df.columns]
        return df.dropna(how="all")

    def _parse_date(self, value) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%m/%d/%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        raise ValueError(f"Não foi possível interpretar a data: {value}")

    def _parse_amount(self, value) -> Decimal:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value)).quantize(Decimal("0.01"))
        if isinstance(value, str):
            cleaned = value.strip().replace("R$", "").replace(" ", "")
            cleaned = cleaned.replace("(", "-").replace(")", "").replace(" ", "")
            has_br_decimal = "," in cleaned
            if has_br_decimal:
                cleaned = cleaned.replace(".", "").replace(",", ".")
            else:
                cleaned = cleaned.replace(",", "")
            return Decimal(cleaned).quantize(Decimal("0.01"))
        raise ValueError(f"Não foi possível interpretar o valor: {value}")
