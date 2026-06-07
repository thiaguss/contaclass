from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from contaclass.models.result import ProcessingBatchResult

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CONFIRMED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CONFIRMED_FONT = Font(color="276221")

REVIEW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
REVIEW_FONT = Font(color="9C6500")

NOT_FOUND_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NOT_FOUND_FONT = Font(color="9C0006")

COLUMNS = [
    ("#", 5),
    ("Data", 14),
    ("Cod Débito", 14),
    ("Cod Crédito", 14),
    ("Valor", 16),
    ("Cód Histórico", 18),
    ("Fornecedor", 35),
    ("Inicia Lote", 14),
    ("Matriz/Filial", 16),
    ("Status", 18),
    ("Score (%)", 12),
    ("Match", 15),
    ("Sugestão Fuzzy", 30),
]

SUMMARY_COLUMNS = [
    ("Métrica", 30),
    ("Valor", 20),
]


class ExcelWriter:
    def __init__(self, threshold: float = 70.0):
        self.threshold = threshold

    def write(
        self,
        batch: ProcessingBatchResult,
        output_path: str | Path,
        client_name: str = "",
        reference_month: str = "",
    ):
        wb = Workbook()

        ws = wb.active
        ws.title = "Classificação"

        self._write_header(ws)
        self._write_data(ws, batch)
        self._auto_width(ws)

        ws_summary = wb.create_sheet("Resumo")
        self._write_summary(ws_summary, batch, client_name, reference_month)

        wb.save(output_path)
        wb.close()

    def _write_header(self, ws):
        for col_idx, (label, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

    def _write_data(self, ws, batch: ProcessingBatchResult):
        for row_idx, entry in enumerate(batch.entries, start=2):
            status = entry.status
            is_confirmed = status == "confirmed"
            is_review = status == "review"
            is_not_found = status == "not_found"

            if is_confirmed:
                fill = CONFIRMED_FILL
                font = CONFIRMED_FONT
            elif is_review:
                fill = REVIEW_FILL
                font = REVIEW_FONT
            else:
                fill = NOT_FOUND_FILL
                font = NOT_FOUND_FONT

            status_label = {
                "confirmed": "Confirmado",
                "review": "Revisar",
                "not_found": "Não Encontrado",
            }.get(status, status)

            values = [
                entry.row_number,
                entry.entry_date,
                entry.debit_code or "",
                entry.credit_code or "",
                float(entry.amount) if entry.amount else 0.0,
                entry.codigo_historico or "",
                entry.raw_supplier,
                entry.inicia_lote or "",
                entry.codigo_matriz_filial or "",
                status_label,
                entry.confidence_score,
                entry.match_type or "",
                entry.matched_supplier or "",
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(vertical="center")

    def _write_summary(
        self, ws, batch: ProcessingBatchResult, client_name: str, reference_month: str
    ):
        for col_idx, (label, _) in enumerate(SUMMARY_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        summary_data = [
            ("Cliente", client_name),
            ("Mês de Referência", reference_month),
            ("Data de Processamento", date.today().isoformat()),
            ("Total de Lançamentos", batch.total_entries),
            ("Confirmados", batch.confirmed_count),
            ("Revisar", batch.review_count),
            ("Não Encontrados", batch.not_found_count),
            ("Taxa de Automação (%)", batch.automation_rate),
            ("Tempo de Processamento (ms)", batch.processing_time_ms),
        ]

        for row_idx, (metric, value) in enumerate(summary_data, start=2):
            ws.cell(row=row_idx, column=1, value=metric)
            ws.cell(row=row_idx, column=2, value=value)

        self._auto_width(ws)

    def _auto_width(self, ws):
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                val = str(cell.value or "")
                max_length = max(max_length, len(val))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)
