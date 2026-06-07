from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from contaclass.models.entry import HistoricalEntry, NewEntry
from contaclass.io.excel_reader import ExcelReader


@pytest.fixture
def hist_xlsx(tmp_path):
    """Cria um arquivo Excel histórico de exemplo com múltiplas abas."""
    path = tmp_path / "historico.xlsx"
    wb = openpyxl.Workbook()

    # Aba 1: Banco do Brasil
    ws1 = wb.active
    ws1.title = "Banco do Brasil"
    ws1.append(["Data", "Cod Débito", "Cod Crédito", "Valor", "Fornecedor"])
    ws1.append(["01/01/2026", "503", "101", 1500.00, "VIVO"])
    ws1.append(["15/01/2026", "504", "102", 850.00, "CEMIG"])
    ws1.append(["20/01/2026", "505", "103", 2200.00, "EQUATORIAL"])

    # Aba 2: Santander
    ws2 = wb.create_sheet("Santander")
    ws2.append(["Data", "Cod Débito", "Cod Crédito", "Valor", "Fornecedor"])
    ws2.append(["05/01/2026", "506", "104", 3000.00, "TIM"])
    ws2.append(["10/01/2026", "507", "105", 1200.00, "CLARO"])

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def novo_xlsx(tmp_path):
    """Cria um arquivo Excel novo de exemplo."""
    path = tmp_path / "novo.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws.append(["Data", "Valor", "Fornecedor"])
    ws.append(["01/02/2026", 1500.00, "VIVO"])
    ws.append(["05/02/2026", 900.00, "CEMIG"])
    ws.append(["10/02/2026", 500.00, "EMPRESA NOVA XYZ"])
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def novo_csv(tmp_path):
    path = tmp_path / "novo.csv"
    path.write_text("Data,Valor,Fornecedor\n01/02/2026,1500.00,VIVO\n05/02/2026,900.00,CEMIG\n", encoding="utf-8")
    return path


def test_list_tabs(hist_xlsx):
    reader = ExcelReader(hist_xlsx)
    tabs = reader.list_tabs()
    assert len(tabs) == 2
    names = [t["name"] for t in tabs]
    assert "Banco do Brasil" in names
    assert "Santander" in names


def test_read_historical_all_tabs(hist_xlsx):
    reader = ExcelReader(hist_xlsx)
    entries = reader.read_historical()
    assert len(entries) == 5
    assert all(isinstance(e, HistoricalEntry) for e in entries)


def test_read_historical_filtered_tab(hist_xlsx):
    reader = ExcelReader(hist_xlsx)
    entries = reader.read_historical(tab_filter=["Santander"])
    assert len(entries) == 2
    assert all(e.tab_name == "Santander" for e in entries)


def test_read_historical_detects_columns(hist_xlsx):
    reader = ExcelReader(hist_xlsx)
    entries = reader.read_historical()
    entry = entries[0]
    assert entry.debit_code == "503"
    assert entry.credit_code == "101"
    assert entry.raw_supplier == "VIVO"
    assert entry.amount == 1500.00


def test_read_new(novo_xlsx):
    reader = ExcelReader(novo_xlsx)
    entries = reader.read_new()
    assert len(entries) == 3
    assert all(isinstance(e, NewEntry) for e in entries)
    assert entries[0].raw_supplier == "VIVO"


def test_read_new_amount_precision(novo_xlsx):
    reader = ExcelReader(novo_xlsx)
    entries = reader.read_new()
    assert entries[0].amount == 1500.00


def test_read_new_date(novo_xlsx):
    reader = ExcelReader(novo_xlsx)
    entries = reader.read_new()
    assert entries[0].entry_date == date(2026, 2, 1)


def test_preview_new(novo_xlsx):
    reader = ExcelReader(novo_xlsx)
    preview = reader.preview_new(n_rows=2)
    assert len(preview) == 2
    assert "supplier" in preview[0]


def test_read_historical_column_map(hist_xlsx):
    reader = ExcelReader(hist_xlsx)
    manual_map = {
        "entry_date": "Data",
        "debit_code": "Cod Débito",
        "credit_code": "Cod Crédito",
        "amount": "Valor",
        "supplier": "Fornecedor",
    }
    entries = reader.read_historical(column_map=manual_map)
    assert len(entries) == 5


def test_read_new_column_map(novo_xlsx):
    reader = ExcelReader(novo_xlsx)
    manual_map = {
        "entry_date": "Data",
        "amount": "Valor",
        "supplier": "Fornecedor",
    }
    entries = reader.read_new(column_map=manual_map)
    assert len(entries) == 3


def test_empty_sheet_skipped(tmp_path):
    path = tmp_path / "vazio.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data", "Valor", "Fornecedor"])
    wb.save(path)
    wb.close()
    reader = ExcelReader(path)
    entries = reader.read_new()
    assert len(entries) == 0


def test_parse_br_amount():
    reader = ExcelReader("dummy.xlsx")
    assert reader._parse_amount("1.500,00") == Decimal("1500.00")
    assert reader._parse_amount("1500,00") == Decimal("1500.00")
    assert reader._parse_amount("1.000.000,00") == Decimal("1000000.00")
    assert reader._parse_amount("0,00") == Decimal("0.00")
    assert reader._parse_amount("R$ 1.500,00") == Decimal("1500.00")
    assert reader._parse_amount("(1.500,00)") == Decimal("-1500.00")
    assert reader._parse_amount("-1500,00") == Decimal("-1500.00")


def test_parse_br_amount_float():
    reader = ExcelReader("dummy.xlsx")
    assert reader._parse_amount(1500.50) == Decimal("1500.50")
    assert reader._parse_amount(1500) == Decimal("1500.00")


# ─── Fixtures para schema fixo de 8 colunas ────────────────────────────────


@pytest.fixture
def hist_8col_xlsx(tmp_path):
    """Arquivo histórico com layout real de 8 colunas."""
    path = tmp_path / "historico_8col.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Banco do Brasil"
    ws.append(["Data", "cód conta débito", "cód Conta Credito", "Valor",
               "CódHistórico", "Complemento Histórico", "inicia Lote", "Código Matriz/Filial"])
    ws.append(["01/01/2026", "503", "101", 1500.00, "H001", "VIVO", "S", "1"])
    ws.append(["15/01/2026", "504", "102", 850.00, "H002", "CEMIG", "N", "1"])
    ws.append(["20/01/2026", "505", "103", 2200.00, "H003", "EQUATORIAL", "S", "2"])
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def novo_8col_xlsx(tmp_path):
    """Arquivo novo com layout real de 8 colunas."""
    path = tmp_path / "novo_8col.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    ws.append(["Data", "cód conta débito", "cód Conta Credito", "Valor",
               "CódHistórico", "Complemento Histórico", "inicia Lote", "Código Matriz/Filial"])
    ws.append(["01/02/2026", "", "", 1500.00, "H010", "VIVO", "S", "1"])
    ws.append(["05/02/2026", "", "", 900.00, "H011", "CEMIG", "N", "1"])
    wb.save(path)
    wb.close()
    return path


# ─── Testes de detecção por schema fixo ─────────────────────────────────────


def _check_8col_schema_detected(reader, path):
    """Verifica se o schema fixo foi detectado chamando _detect_by_fixed_schema diretamente."""
    import pandas as pd
    xls = pd.ExcelFile(path)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], header=0)
    clean = reader._clean_dataframe(df)
    mapping = reader._detect_by_fixed_schema(list(clean.columns))
    xls.close()
    return mapping


def test_fixed_schema_detection_historical(hist_8col_xlsx):
    reader = ExcelReader(hist_8col_xlsx)
    entries = reader.read_historical()
    assert len(entries) == 3
    assert all(isinstance(e, HistoricalEntry) for e in entries)


def test_fixed_schema_populates_extra_fields_historical(hist_8col_xlsx):
    reader = ExcelReader(hist_8col_xlsx)
    entries = reader.read_historical()
    entry = entries[0]
    assert entry.debit_code == "503"
    assert entry.credit_code == "101"
    assert entry.raw_supplier == "VIVO"
    assert entry.amount == Decimal("1500.00")
    assert entry.codigo_historico == "H001"
    assert entry.codigo_matriz_filial == "1"
    assert entry.inicia_lote == "S"


def test_fixed_schema_detection_new(novo_8col_xlsx):
    reader = ExcelReader(novo_8col_xlsx)
    entries = reader.read_new()
    assert len(entries) == 2
    assert all(isinstance(e, NewEntry) for e in entries)


def test_fixed_schema_populates_extra_fields_new(novo_8col_xlsx):
    reader = ExcelReader(novo_8col_xlsx)
    entries = reader.read_new()
    entry = entries[0]
    assert entry.raw_supplier == "VIVO"
    assert entry.amount == Decimal("1500.00")
    assert entry.codigo_historico == "H010"
    assert entry.codigo_matriz_filial == "1"
    assert entry.inicia_lote == "S"


def test_fixed_schema_all_entries_extra_fields(hist_8col_xlsx):
    reader = ExcelReader(hist_8col_xlsx)
    entries = reader.read_historical()
    for e in entries:
        assert e.codigo_historico is not None
        assert e.codigo_matriz_filial is not None
        assert e.inicia_lote is not None


def test_fixed_schema_fallback_to_aliases(hist_xlsx):
    """Arquivo legado de 5 colunas ainda funciona via aliases."""
    reader = ExcelReader(hist_xlsx)
    entries = reader.read_historical()
    assert len(entries) == 5
    assert entries[0].codigo_historico is None
    assert entries[0].codigo_matriz_filial is None


def test_detect_by_fixed_schema_returns_correct_mapping(hist_8col_xlsx):
    reader = ExcelReader(hist_8col_xlsx)
    mapping = _check_8col_schema_detected(reader, hist_8col_xlsx)
    assert mapping is not None
    assert mapping["entry_date"] == "Data"
    assert mapping["debit_code"] == "cód conta débito"
    assert mapping["credit_code"] == "cód Conta Credito"
    assert mapping["amount"] == "Valor"
    assert mapping["codigo_historico"] == "CódHistórico"
    assert mapping["supplier"] == "Complemento Histórico"
    assert mapping["inicia_lote"] == "inicia Lote"
    assert mapping["codigo_matriz_filial"] == "Código Matriz/Filial"


def test_detect_by_fixed_schema_rejects_non_matching(tmp_path):
    """Schema com menos de 6/8 match deve retornar None."""
    path = tmp_path / "incompativel.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D", "E", "F", "G", "H"])
    ws.append(["1", "2", "3", "4", "5", "6", "7", "8"])
    wb.save(path)
    wb.close()

    reader = ExcelReader(path)
    import pandas as pd
    xls = pd.ExcelFile(path)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], header=0)
    clean = reader._clean_dataframe(df)
    mapping = reader._detect_by_fixed_schema(list(clean.columns))
    xls.close()
    assert mapping is None
