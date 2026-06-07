from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from contaclass.io.excel_writer import ExcelWriter
from contaclass.models.result import MatchResult, ProcessingBatchResult


@pytest.fixture
def sample_batch():
    batch = ProcessingBatchResult()
    batch.total_entries = 3
    batch.processing_time_ms = 1250

    batch.entries = [
        MatchResult(
            row_number=1,
            entry_date=date(2026, 2, 1),
            raw_supplier="VIVO",
            normalized_supplier="VIVO",
            amount=Decimal("1500.00"),
            debit_code="503",
            credit_code="101",
            confidence_score=100.0,
            status="confirmed",
            match_type="exact",
            matched_supplier="VIVO",
            matched_supplier_normalized="VIVO",
        ),
        MatchResult(
            row_number=2,
            entry_date=date(2026, 2, 5),
            raw_supplier="TELEFONICA BRASIL",
            normalized_supplier="TELEFONICA BRASIL",
            amount=Decimal("900.00"),
            debit_code="503",
            credit_code="101",
            confidence_score=78.5,
            status="review",
            match_type="fuzzy",
            matched_supplier="VIVO",
            matched_supplier_normalized="VIVO",
        ),
        MatchResult(
            row_number=3,
            entry_date=date(2026, 2, 10),
            raw_supplier="EMPRESA NOVA XYZ",
            normalized_supplier="EMPRESA NOVA XYZ",
            amount=Decimal("500.00"),
            debit_code=None,
            credit_code=None,
            confidence_score=0.0,
            status="not_found",
            match_type=None,
            matched_supplier=None,
            matched_supplier_normalized=None,
        ),
    ]

    batch.confirmed_count = 1
    batch.review_count = 1
    batch.not_found_count = 1

    return batch


def test_excel_written_with_colors(tmp_path, sample_batch):
    output = tmp_path / "resultado.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output, client_name="Empresa ABC", reference_month="2026-02")

    assert output.exists()

    wb = openpyxl.load_workbook(output)
    assert "Classificação" in wb.sheetnames
    assert "Resumo" in wb.sheetnames
    wb.close()


def test_header_formatting(tmp_path, sample_batch):
    output = tmp_path / "resultado.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    header = ws.cell(row=1, column=1)
    assert header.value == "#"
    assert header.font.bold is True
    wb.close()


def test_confirmed_row_green(tmp_path, sample_batch):
    output = tmp_path / "resultado.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    cell = ws.cell(row=2, column=3)
    assert cell.value == "VIVO"
    wb.close()


def test_review_row_yellow(tmp_path, sample_batch):
    output = tmp_path / "resultado.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    cell = ws.cell(row=3, column=3)
    assert cell.value == "TELEFONICA BRASIL"
    wb.close()


def test_not_found_row_red(tmp_path, sample_batch):
    output = tmp_path / "resultado.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    cell = ws.cell(row=4, column=3)
    assert cell.value == "EMPRESA NOVA XYZ"
    wb.close()


def test_summary_sheet(tmp_path, sample_batch):
    output = tmp_path / "resultado.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output, client_name="ABC", reference_month="2026-02")

    wb = openpyxl.load_workbook(output)
    ws = wb["Resumo"]
    values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(2, 11)}
    assert values["Cliente"] == "ABC"
    assert values["Total de Lançamentos"] == 3
    assert values["Confirmados"] == 1
    assert values["Taxa de Automação (%)"] == pytest.approx(33.33, rel=0.1)
    wb.close()


def test_multiple_writes_different_files(tmp_path, sample_batch):
    writer = ExcelWriter()
    out1 = tmp_path / "r1.xlsx"
    out2 = tmp_path / "r2.xlsx"
    writer.write(sample_batch, out1)
    writer.write(sample_batch, out2)
    assert out1.exists()
    assert out2.exists()


@pytest.fixture
def batch_with_extra_fields():
    batch = ProcessingBatchResult()
    batch.total_entries = 1
    batch.processing_time_ms = 500
    batch.entries = [
        MatchResult(
            row_number=1,
            entry_date=date(2026, 6, 1),
            raw_supplier="VIVO",
            normalized_supplier="VIVO",
            amount=Decimal("1500.00"),
            debit_code="503",
            credit_code="101",
            confidence_score=100.0,
            status="confirmed",
            match_type="exact",
            matched_supplier="VIVO",
            matched_supplier_normalized="VIVO",
            codigo_historico="H001",
            codigo_matriz_filial="1",
            inicia_lote="S",
        ),
    ]
    batch.confirmed_count = 1
    batch.review_count = 0
    batch.not_found_count = 0
    return batch


def test_extra_columns_present_in_header(tmp_path, batch_with_extra_fields):
    output = tmp_path / "extra.xlsx"
    writer = ExcelWriter()
    writer.write(batch_with_extra_fields, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
    assert "Cód Histórico" in headers
    assert "Matriz/Filial" in headers
    assert "Inicia Lote" in headers
    wb.close()


def test_extra_columns_values_preserved(tmp_path, batch_with_extra_fields):
    output = tmp_path / "extra.xlsx"
    writer = ExcelWriter()
    writer.write(batch_with_extra_fields, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    assert ws.cell(row=2, column=7).value == "H001"
    assert ws.cell(row=2, column=8).value == "1"
    assert ws.cell(row=2, column=9).value == "S"
    wb.close()


def test_extra_columns_empty_when_not_provided(tmp_path, sample_batch):
    output = tmp_path / "sem_extra.xlsx"
    writer = ExcelWriter()
    writer.write(sample_batch, output)

    wb = openpyxl.load_workbook(output)
    ws = wb["Classificação"]
    assert ws.cell(row=2, column=7).value is None
    assert ws.cell(row=2, column=8).value is None
    assert ws.cell(row=2, column=9).value is None
    wb.close()
